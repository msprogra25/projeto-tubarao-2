"""Geração da planilha de vendas em Excel usando openpyxl."""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

STORAGE_DIR = os.path.join(os.path.dirname(__file__), "storage", "excel")
os.makedirs(STORAGE_DIR, exist_ok=True)

VERDE_FILL = PatternFill(start_color="1DB954", end_color="1DB954", fill_type="solid")


def gerar_planilha_vendas(vendas: list) -> tuple[str, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendas"

    cabecalho = ["Produto", "Valor Unitário (R$)", "Quantidade", "Subtotal (R$)", "Data/Hora da Venda"]
    ws.append(cabecalho)
    for col in range(1, len(cabecalho) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = VERDE_FILL
        cell.alignment = Alignment(horizontal="center")

    for venda in vendas:
        ws.append([
            venda["nome"],
            venda["valor_unitario"],
            venda["quantidade"],
            venda["subtotal"],
            venda["criado_em"],
        ])

    for i, largura in enumerate([30, 20, 12, 16, 22], start=1):
        ws.column_dimensions[chr(64 + i)].width = largura

    filename = f"relatorio_vendas_{int(datetime.now().timestamp())}.xlsx"
    filepath = os.path.join(STORAGE_DIR, filename)
    wb.save(filepath)
    return filepath, filename
